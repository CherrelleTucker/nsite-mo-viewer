"""
Map contact solution names to solution_ids

This script:
1. Reads contacts from MO-DB_Contacts.xlsx
2. Reads solutions from MO-DB_Solutions.xlsx
3. Attempts auto-matching using keywords
4. Outputs an Excel file with contact_id and solution_id for copying

Usage:
1. Run: python map_contact_solutions.py
2. Review the output mapping file
3. Copy solution_id column to MO-DB_Contacts
"""

import re
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl


def load_contacts(filepath):
    """Load all contacts from Excel file"""
    contacts = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        headers = [h.lower() if h else '' for h in headers]

        # Find column indices
        contact_id_idx = headers.index('contact_id') if 'contact_id' in headers else None
        solution_idx = headers.index('solution') if 'solution' in headers else None
        name_idx = headers.index('name') if 'name' in headers else None

        if contact_id_idx is None:
            print("Warning: 'contact_id' column not found")
            return []
        if solution_idx is None:
            print("Warning: 'solution' column not found")
            return []

        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            contact_id = row[contact_id_idx] if contact_id_idx < len(row) else None
            solution = row[solution_idx] if solution_idx < len(row) else None
            name = row[name_idx] if name_idx is not None and name_idx < len(row) else None

            if contact_id:
                contacts.append({
                    'contact_id': str(contact_id).strip(),
                    'solution': str(solution).strip() if solution else '',
                    'name': str(name).strip() if name else ''
                })

        wb.close()
    except FileNotFoundError:
        print(f"File not found: {filepath}")
        return []
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        return []
    return contacts


def load_solutions_db(filepath):
    """Load solutions from MO-DB_Solutions Excel file"""
    solutions = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Get headers
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        headers = [h.lower() if h else '' for h in headers]

        # Find column indices
        sol_id_idx = headers.index('solution_id') if 'solution_id' in headers else None
        name_idx = headers.index('name') if 'name' in headers else None
        title_idx = headers.index('full_title') if 'full_title' in headers else None

        if sol_id_idx is None:
            print("Warning: 'solution_id' column not found in solutions")
            return []

        for row in ws.iter_rows(min_row=2, values_only=True):
            sol_id = row[sol_id_idx] if sol_id_idx < len(row) else None
            name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
            title = row[title_idx] if title_idx is not None and title_idx < len(row) else None

            if sol_id:
                solutions.append({
                    'solution_id': str(sol_id).strip(),
                    'name': str(name).strip() if name else '',
                    'full_title': str(title).strip() if title else '',
                })

        wb.close()
    except FileNotFoundError:
        print(f"File not found: {filepath}")
        return []
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        return []
    return solutions

def extract_keywords(text):
    """Extract meaningful keywords from a solution name"""
    # Remove common words
    stopwords = {'to', 'the', 'a', 'an', 'and', 'or', 'for', 'of', 'in', 'on', 'at', 'by', 'from', 'with', 'data', 'access', 'product', 'products'}

    # Clean and tokenize
    text = text.lower()
    text = re.sub(r'[^\w\s]', ' ', text)
    words = text.split()

    # Filter
    keywords = [w for w in words if w not in stopwords and len(w) > 2]
    return set(keywords)

def find_best_match(contact_solution, solutions_db):
    """Try to find the best matching solution_id for a contact solution name"""
    contact_keywords = extract_keywords(contact_solution)
    contact_lower = contact_solution.lower()

    best_match = None
    best_score = 0

    for sol in solutions_db:
        score = 0
        sol_id = sol['solution_id'].lower()
        sol_name = sol['name'].lower()
        sol_title = sol['full_title'].lower()

        # Exact match on name or title
        if contact_lower == sol_name or contact_lower == sol_title:
            return sol['solution_id'], 100

        # Check if solution_id appears in contact solution
        if sol_id.replace('_', ' ') in contact_lower or sol_id.replace('_', '') in contact_lower.replace(' ', ''):
            score += 50

        # Check keyword overlap
        sol_keywords = extract_keywords(sol['name'] + ' ' + sol['full_title'] + ' ' + sol['solution_id'])
        overlap = contact_keywords & sol_keywords
        score += len(overlap) * 10

        # Check for acronym matches (e.g., HLS, OPERA, EMIT)
        acronyms = re.findall(r'\b[A-Z]{2,}\b', contact_solution)
        for acr in acronyms:
            if acr.lower() in sol_id or acr.lower() in sol_name:
                score += 30

        if score > best_score:
            best_score = score
            best_match = sol['solution_id']

    return best_match, best_score

def main():
    import os

    # File paths - relative to script location
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)  # nsite-mo-viewer
    db_dir = os.path.join(os.path.dirname(project_dir), 'database-files', 'MO-Viewer Databases')

    contacts_file = os.path.join(db_dir, 'MO-DB_Contacts.xlsx')
    solutions_file = os.path.join(db_dir, 'MO-DB_Solutions.xlsx')
    output_file = os.path.join(db_dir, 'contact_solution_id_mapping.xlsx')

    print(f"Contacts file: {contacts_file}")
    print(f"Solutions file: {solutions_file}")

    print("\nLoading contacts...")
    contacts = load_contacts(contacts_file)
    print(f"Found {len(contacts)} contacts")

    print("Loading solutions database...")
    solutions_db = load_solutions_db(solutions_file)
    print(f"Found {len(solutions_db)} solutions")

    if not contacts or not solutions_db:
        print("\nError: Could not load data files")
        return

    # Build a mapping from solution name to solution_id
    print("\nBuilding solution name -> solution_id mapping...")

    # First, count unique solution names in contacts
    solution_names = defaultdict(int)
    for c in contacts:
        if c['solution']:
            solution_names[c['solution']] += 1

    print(f"Found {len(solution_names)} unique solution values in contacts")

    # Match each unique solution name to a solution_id
    name_to_id = {}
    for sol_name in solution_names:
        match, score = find_best_match(sol_name, solutions_db)
        name_to_id[sol_name] = {
            'solution_id': match or '',
            'confidence': score
        }

    # Now create output with every contact row
    print("\nMatching contacts to solution_ids...")
    output_rows = []

    for c in contacts:
        sol_name = c['solution']
        if sol_name and sol_name in name_to_id:
            mapped = name_to_id[sol_name]
            output_rows.append({
                'contact_id': c['contact_id'],
                'name': c['name'],
                'solution': sol_name,
                'solution_id': mapped['solution_id'],
                'confidence': mapped['confidence']
            })
        else:
            output_rows.append({
                'contact_id': c['contact_id'],
                'name': c['name'],
                'solution': sol_name,
                'solution_id': '',
                'confidence': 0
            })

    # Write output Excel file
    print(f"\nWriting output to {output_file}...")
    wb = openpyxl.Workbook()

    # Sheet 1: All contacts with mapped solution_id
    ws1 = wb.active
    ws1.title = 'Contact Mappings'
    ws1.append(['contact_id', 'name', 'solution', 'solution_id', 'confidence'])
    for row in output_rows:
        ws1.append([row['contact_id'], row['name'], row['solution'], row['solution_id'], row['confidence']])

    # Sheet 2: Mapping reference (unique solution names)
    ws2 = wb.create_sheet('Solution Name Reference')
    ws2.append(['solution_name', 'contact_count', 'suggested_solution_id', 'confidence'])
    for sol_name, count in sorted(solution_names.items(), key=lambda x: -x[1]):
        mapped = name_to_id.get(sol_name, {'solution_id': '', 'confidence': 0})
        ws2.append([sol_name, count, mapped['solution_id'], mapped['confidence']])

    # Sheet 3: Available solutions for reference
    ws3 = wb.create_sheet('Available Solutions')
    ws3.append(['solution_id', 'name', 'full_title'])
    for sol in solutions_db:
        ws3.append([sol['solution_id'], sol['name'], sol['full_title']])

    wb.save(output_file)
    print(f"Saved to: {output_file}")

    # Summary
    matched = sum(1 for r in output_rows if r['solution_id'])
    high_conf = sum(1 for r in output_rows if r['confidence'] >= 50)
    low_conf = sum(1 for r in output_rows if 0 < r['confidence'] < 50)
    no_match = sum(1 for r in output_rows if r['confidence'] == 0 and r['solution'])
    no_solution = sum(1 for r in output_rows if not r['solution'])

    print(f"\nSummary ({len(contacts)} contacts):")
    print(f"  High confidence matches (>=50): {high_conf}")
    print(f"  Low confidence matches (<50):   {low_conf}")
    print(f"  No match found:                 {no_match}")
    print(f"  No solution value:              {no_solution}")
    print(f"\nOutput file has 3 sheets:")
    print(f"  1. Contact Mappings - Every contact with solution_id (copy this column)")
    print(f"  2. Solution Name Reference - Unique solution names for review")
    print(f"  3. Available Solutions - All solution_ids for manual lookup")

if __name__ == '__main__':
    main()
